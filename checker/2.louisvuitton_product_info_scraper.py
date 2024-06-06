import os
import json
import re
import requests
from datetime import datetime
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from colorama import Fore, Style
import time
import textwrap


def create_initial_file():
    source_filename = '루이비통.xlsx'
    current_dir = os.getcwd()
    source_path = os.path.join(current_dir, source_filename)

    if os.path.exists(source_path):
        print("루이비통.xlsx 파일이 이미 존재합니다. 계속 진행합니다.")
        return

    workbook = Workbook()
    sheet = workbook.active
    columns_titles = ['품번', '매장가', '공홈여부', '재고현황', '재고매장', '제품명', '소재', '사이즈', '특징', '세부설명']

    for col, title in enumerate(columns_titles, start=1):
        sheet.cell(row=1, column=col, value=title)

    workbook.save(filename=source_path)
    print("루이비통.xlsx 파일을 생성하였습니다. 생성된 파일에 품번을 입력하고 Enter를 눌러 계속 진행하세요.")
    input()


def setup_driver():
    options = Options()
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument('--disable-infobars')
    # options.add_argument('--headless')
    options.add_argument('--start-maximized')  # 창 최대화
    options.add_argument(
        'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3')
    options.add_argument('--log-level=3')

    service = ChromeService(ChromeDriverManager().install())
    service.log_path = os.devnull
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(5)
    return driver


def save_progress(progress_file, sheet_name, row):
    try:
        with open(progress_file, 'w') as file:
            file.write(f"{sheet_name},{row}")
    except Exception as e:
        print(f"Failed to save progress: {e}")
    print(f"Saved progress to {progress_file}: {sheet_name},{row}")


def load_progress(progress_file):
    if os.path.exists(progress_file):
        with open(progress_file, 'r') as file:
            content = file.read().strip()
            if content:
                sheet_name, row = content.split(',')
                return sheet_name, int(row)
            else:
                return None, 0
    else:
        return None, 0


def search_product(driver, product_code):
    try:
        search_button = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".lv-header__utility-label.list-label-s"))
        )
        search_button.click()
        time.sleep(2)  # 검색창이 나타날 때까지 대기

        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#searchHeaderInput"))
        )
        search_input.clear()  # 검색창 초기화
        search_input.send_keys(product_code)  # 품번 입력
        search_input.send_keys(Keys.ENTER)  # Enter 키 입력
        time.sleep(5)  # 검색 결과가 로딩될 때까지 대기

        # 팝업 닫기
        close_popup(driver)
    except Exception as e:
        print(f"검색 중 오류 발생: {e}")


def close_popup(driver):
    try:
        close_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, ".lv-notifications__close.lv-button.-only-icon"))
        )
        close_button.click()
        print("Popup closed")
    except Exception as e:
        print("No popup found to close")


def get_product_info(driver, product_code):
    try:
        search_product(driver, product_code)

        if "product not found" in driver.page_source.lower():
            raise ValueError("Product not found")

        try:
            price_element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.CLASS_NAME, 'lv-product__price'))
            )
            price_text = price_element.text.strip()
            price_text = re.sub(r'[₩\s]', '', price_text)
        except Exception as e:
            print(f"가격 정보를 불러오는 중 예외 발생: {e}")
            price_text = "N/A"

        try:
            product_name = driver.find_element(By.CLASS_NAME, 'lv-product__name').text
        except:
            product_name = "N/A"

        try:
            material_button = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "button.lv-product-variation-selector"))
            )
            material_title = material_button.find_element(By.XPATH, ".//span[contains(text(), '소재')]")
            material_value = material_button.find_element(By.CSS_SELECTOR, '.lv-product-variation-selector__value').text
            material = material_value.strip()
        except Exception as e:
            print(f"소재 정보 파싱 중 오류 발생: {e}")
            material = "N/A"

        try:
            details_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '.lv-product-features-buttons__button'))
            )
            details_button.click()
            description_text = WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, '#modalContent p'))
            ).text
        except:
            description_text = "N/A"

        try:
            size_element = driver.find_element(By.CSS_SELECTOR,
                                               '#modalContent > div.lv-product-detailed-features > div > div.lv-product-dimension.body-s.lv-product-detailed-features__dimensions')
            if size_element:
                size_text = size_element.text.strip()
                size_text = ' '.join(size_text.splitlines())
            else:
                size_text = "N/A"
        except Exception as e:
            size_text = "N/A"

        try:
            features = driver.find_elements(By.CSS_SELECTOR, '.lv-product-detailed-features__description li')
            feature_text = "\n".join([feature.text.strip() for feature in features if feature.text.strip() != ""])
        except Exception as e:
            feature_text = "N/A"

        return price_text, product_name, material, size_text, feature_text, description_text

    except Exception as e:
        print(f"제품 {product_code}에 대한 오류 발생: {str(e)}")
        return "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"


def write_to_cell(sheet, column, row_num, value):
    if value is None or value == '' or value == '0':
        value = 'N/A'

    value = ''.join(char for char in value if char.isprintable())
    sheet.cell(row=row_num, column=column, value=value)


def main_task():
    create_initial_file()

    current_dir = os.getcwd()
    datetime_now = datetime.now().strftime('%Y%m%d')
    source_filename = '루이비통.xlsx'
    target_filename = f'루이비통_{datetime_now}.xlsx'
    progress_file_name = f'progress_{datetime_now}.txt'

    source_path = os.path.join(current_dir, source_filename)
    target_path = os.path.join(current_dir, target_filename)
    progress_file = os.path.join(current_dir, progress_file_name)

    # 사용자 정의 상태 변수
    ON_WEBSITE_SOLD = "O"  # 공홈에도 있고 실제로 판매 중인 제품
    NOT_ON_WEBSITE_SOLD = "X"  # 공홈에는 없지만 실제로 판매 중인 제품
    NOT_AVAILABLE = "N/A"  # 공홈에도 없고 실제로 판매하지 않는 제품
    ON_WEBSITE_NOT_SOLD = "O"  # 공홈에는 있지만 판매하지 않는 제품
    max_stores_to_save = 5

    workbook = load_workbook(filename=source_path) if os.path.exists(source_path) else Workbook()

    if os.path.exists(target_path):
        workbook = load_workbook(filename=target_path)
    else:
        workbook = Workbook()
        workbook.remove(workbook.active)
        source_workbook = load_workbook(filename=source_path)
        for sheet_name in source_workbook.sheetnames:
            source_sheet = source_workbook[sheet_name]
            target_sheet = workbook.create_sheet(sheet_name)
            for row in source_sheet.iter_rows(values_only=True):
                target_sheet.append(row)

    columns_titles = ['품번', '매장가', '공홈여부', '재고현황', '재고매장', '제품명', '소재', '사이즈', '특징', '세부설명']
    for sheet in workbook.worksheets:
        for col, title in enumerate(columns_titles, start=1):
            sheet.cell(row=1, column=col, value=title)

    workbook.save(filename=target_path)
    last_processed_sheet, last_processed_row = load_progress(progress_file)

    start_row = 2

    if last_processed_sheet is None:
        last_processed_sheet = workbook.sheetnames[0]
    if last_processed_row is None or last_processed_row < start_row:
        last_processed_row = start_row

    driver = setup_driver()
    driver.get("https://kr.louisvuitton.com/")

    # Close popup if it appears
    close_popup(driver)

    sheet_found = False
    for sheet_name in workbook.sheetnames:
        if sheet_name == last_processed_sheet:
            sheet_found = True
        if not sheet_found:
            continue

        sheet = workbook[sheet_name]
        last_row = sheet.max_row

        if sheet_name == last_processed_sheet:
            start_row = last_processed_row
        else:
            start_row = 2

        for row_num in range(start_row, last_row + 1):
            row_data = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
            if not row_data or not row_data[0]:
                break
            product_code = row_data[0][0]
            if product_code is None:
                break

            product_price, product_name, material, size, features, description = get_product_info(driver, product_code)

            total_stores_count = 0
            seoul_dosan_count = 0
            stock_info = []

            if product_price != 'N/A':
                try:
                    url = 'https://api.louisvuitton.com/eco-eu/search-merch-eapi/v1/kor-kr/stores/query'
                    headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36',
                        'Content-Type': 'application/json',
                        'Accept': 'application/json, text/plain, */*',
                        'client_secret': '60bbcdcD722D411B88cBb72C8246a22F',
                        'client_id': '607e3016889f431fb8020693311016c9',
                        'Origin': 'https://kr.louisvuitton.com',
                        'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
                    }
                    data = {
                        "flagShip": False,
                        "country": "KR",
                        "query": "서울",
                        "latitudeA": "37.768825744921735",
                        "latitudeB": "37.36061764438869",
                        "latitudeCenter": "37.56472169465521",
                        "longitudeA": "126.51683339179687",
                        "longitudeB": "127.43144520820312",
                        "longitudeCenter": "126.97413929999999",
                        "query": "",
                        "clickAndCollect": False,
                        "skuId": product_code,
                        "pageType": "productsheet"
                    }
                    response = requests.post(url, headers=headers, data=json.dumps(data))

                    if response.status_code == 200:
                        response_data = response.json()
                        stores_with_stock = [
                            store for store in response_data.get('hits', [])
                            if any(
                                prop.get('value') == 'true' and prop.get('name') == 'stockAvailability'
                                for prop in store.get('additionalProperty', [])
                            )
                        ]
                        total_stores_count = len(stores_with_stock)

                        for store in stores_with_stock:
                            store_name = store.get('name')
                            stock_info.append(store_name)

                        if '루이 비통 메종 서울' in store_name or '루이 비통 서울 도산' in store_name:
                            seoul_dosan_count += 1
                    else:
                        print(f"API 요청이 실패했습니다. 상태 코드: {response.status_code}")
                except Exception as e:
                    print(f"재고 조회 중 오류 발생: {e}")
                    continue

            if total_stores_count > 0:
                if max_stores_to_save > 0 and len(stock_info) > max_stores_to_save:
                    stock_info_to_save = stock_info[:max_stores_to_save]
                else:
                    stock_info_to_save = stock_info
                stock_info_str = '\n'.join(stock_info_to_save)
            else:
                stock_info_str = 'N/A'

            if total_stores_count == 0 and seoul_dosan_count == 0:
                stock_status = 'N/A'
            else:
                stock_status = f"{total_stores_count}({seoul_dosan_count})"

            if product_price == 'N/A' and stock_info_str == 'N/A':
                website_status = NOT_AVAILABLE
            elif product_price != 'N/A' and stock_info_str == 'N/A':
                website_status = ON_WEBSITE_NOT_SOLD
            elif product_price == 'N/A' and stock_info_str != 'N/A':
                website_status = NOT_ON_WEBSITE_SOLD
            else:
                website_status = ON_WEBSITE_SOLD

            write_to_cell(sheet, 2, row_num, product_price)
            write_to_cell(sheet, 3, row_num, website_status)
            write_to_cell(sheet, 4, row_num, stock_status)
            write_to_cell(sheet, 5, row_num, stock_info_str)
            write_to_cell(sheet, 6, row_num, product_name)
            write_to_cell(sheet, 7, row_num, material)
            write_to_cell(sheet, 8, row_num, size)
            write_to_cell(sheet, 9, row_num, features)
            write_to_cell(sheet, 10, row_num, description)

            def print_product_info(product_code, product_price, website_status, stock_status, stock_info,
                                   product_name, material, size, features, description):
                def format_text_block(text, width=70):
                    return textwrap.fill(text, width=width)

                formatted_description = format_text_block(description)
                formatted_features = format_text_block(features)
                formatted_stock_info = "\n".join(stock_info)

                print(Fore.LIGHTCYAN_EX + "[제품 정보]:" + Style.RESET_ALL)
                print(Fore.LIGHTMAGENTA_EX + f"품번: {product_code}" + Style.RESET_ALL)
                print(Fore.LIGHTGREEN_EX + f"매장가: {product_price}" + Style.RESET_ALL)
                print(Fore.LIGHTYELLOW_EX + f"공홈여부: {website_status}" + Style.RESET_ALL)
                print(Fore.LIGHTBLUE_EX + f"재고현황: {stock_status}" + Style.RESET_ALL)
                print(Fore.LIGHTMAGENTA_EX + f"재고매장:\n{formatted_stock_info}" + Style.RESET_ALL)
                print(Fore.LIGHTRED_EX + f"제품명: {product_name}" + Style.RESET_ALL)
                print(Fore.LIGHTCYAN_EX + f"소재: {material}" + Style.RESET_ALL)
                print(Fore.LIGHTGREEN_EX + f"사이즈: {size}" + Style.RESET_ALL)
                print(Fore.LIGHTYELLOW_EX + f"특징:\n{formatted_features}" + Style.RESET_ALL)
                print(Fore.LIGHTBLUE_EX + f"세부 설명:\n{formatted_description}" + Style.RESET_ALL)

            print_product_info(product_code, product_price, website_status, stock_status, stock_info, product_name,
                               material, size, features, description)

            workbook.save(filename=target_path)
            save_progress(progress_file, sheet_name, row_num)
            print(f"제품 정보를 엑셀 파일에 저장했습니다.")

            # 검색 결과를 초기화하기 위해 메인 페이지로 돌아가기
            driver.get("https://kr.louisvuitton.com/")
            time.sleep(5)  # 메인 페이지가 완전히 로드될 때까지 대기

        last_processed_row = start_row
        last_processed_row = 2

    driver.quit()
    print("작업이 완료되었습니다.")


def main():
    while True:
        try:
            main_task()
            break  # 작업이 성공적으로 완료되면 반복을 중단
        except Exception as e:
            print(f"An unexpected error occurred: {e}")
            print("Restarting in 5 seconds...")
            time.sleep(5)

if __name__ == "__main__":
    main()
