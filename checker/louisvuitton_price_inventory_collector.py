import requests
import json
from datetime import datetime
from openpyxl import load_workbook
import os
import time
import sys

# 사용자 정의 상태 변수 확장
ON_WEBSITE_SOLD = "O"  # 공홈에도 있고 실제로 판매 중인 제품
NOT_ON_WEBSITE_SOLD = "X"  # 공홈에는 없지만 실제로 판매 중인 제품
NOT_AVAILABLE = "N/A"  # 공홈에도 없고 실제로 판매하지 않는 제품
ON_WEBSITE_NOT_SOLD = "N/A"  # 공홈에는 있지만 판매하지 않는 제품
max_stores_to_save = 5  # 0이면 매장 전체 출력

current_dir = os.getcwd()
datetime_now = datetime.now().strftime('%Y%m%d')

source_filename = '루이비통.xlsx'
target_filename = f'루이비통_{datetime_now}.xlsx'
progress_file_name = f'progress_{datetime_now}.txt'  # 현재 날짜를 포함한 파일 이름

source_path = os.path.join(current_dir, source_filename)
target_path = os.path.join(current_dir, target_filename)
progress_file = os.path.join(current_dir, progress_file_name)

def get_base_dir():
    if getattr(sys, 'frozen', False):
        return sys._MEIPASS
    else:
        return os.path.dirname(os.path.abspath(__file__))

def save_progress(progress_file, sheet_name, row):
    with open(progress_file, 'w') as file:
        file.write(f"{sheet_name},{row}")

def load_progress(progress_file):
    if os.path.exists(progress_file):
        with open(progress_file, 'r') as file:
            content = file.read()
            if content:
                sheet_name, row = content.split(',')
                return sheet_name, int(row)
    return None, 0

try:
    workbook = load_workbook(filename=source_path)
    workbook.save(filename=target_path)
    print(f"원본 파일을 '{target_filename}'으로 복사했습니다.")
except Exception as e:
    print(f"파일 복사 중 오류: {e}")
    sys.exit()

last_processed_sheet, last_processed_row = load_progress(progress_file)

def get_product_price(product_code):
    # API 요청을 위한 URL 및 헤더 설정
    url = f'https://api.louisvuitton.com/eco-eu/catalog-lvcom/v1/kor-kr/skus/{product_code}/low'
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36',
        'Referer': f'https://kr.louisvuitton.com/kor-kr/products/dopp-kit-monogram-canvas-nvprod1420097v/{product_code}',
        'client_id': '607e3016889f431fb8020693311016c9',
        'client_secret': '60bbcdcD722D411B88cBb72C8246a22F'
    }
    
    # GET 요청을 통해 제품 정보 가져오기
    response = requests.get(url, headers=headers)
    
    # 응답 데이터 파싱
    if response.status_code == 200:
        data = response.json()
        if data.get('skuListSize', 1) == 0:
            return 'N/A'  # 'skuListSize'가 0이면 'N/A' 반환
        for sku in data.get('skuList', []):
            offers = sku.get('offers', {})
            price = offers.get('price', '가격 정보 없음')
            return price  # 첫 번째 제품의 가격을 반환
    else:
        print('응답 상태 코드:', response.status_code)
        print('응답 본문:', response.text)
        return '제품 정보를 가져오는 데 실패하였습니다.'

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    if last_processed_sheet and sheet_name < last_processed_sheet:
        continue
    start_row = last_processed_row + 1 if sheet_name == last_processed_sheet else 2

    # 현재 시트의 전체 행 수 계산
    total_rows = sheet.max_row
    # 처리된 행 수를 저장할 변수 초기화
    processed_rows = 0

    for row in range(start_row, sheet.max_row + 1):
        product_code = sheet.cell(row=row, column=1).value
        if not product_code:
            continue

        # 초기값 설정
        price = 'N/A'
        stock_status = 'N/A'
        website_status = NOT_AVAILABLE  # 초기 상태를 'N/A'로 설정
        stock_info_str = 'N/A'
                        
        # 가격 정보 조회
        try:
            price = get_product_price(product_code)
        except Exception as e:
            print(f"{product_code} 가격 조회 중 오류: {e}")

        # 매장 재고 조회
        total_stores_count = 0
        seoul_dosan_count = 0  # 서울/도산 매장 수 초기화
        stock_info = []  # 재고가 있는 매장 이름을 저장할 리스트

        if price != 'N/A':  # 가격 정보가 'N/A'가 아닌 경우에만 재고 조회를 실행
            try:
                url = 'https://api.louisvuitton.com/eco-eu/search-merch-eapi/v1/kor-kr/stores/query'
                headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36',
                        'Content-Type': 'application/json',
                        'Accept': 'application/json, text/plain, */*',
                        'client_secret': '60bbcdcD722D411B88cBb72C8246a22F',
                        'client_id': '607e3016889f431fb8020693311016c9',
                        'Origin': 'https://kr.louisvuitton.com',
                        'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
                        'Cookie': 'ATG_SESSION_ID=E4i7k6N4uqdHA3we5jdEXtRl.front131-prd; ATGID=anonymous; anonymous_session=true; _dynSessConf=8923089313706214252; JSESSIONID=E4i7k6N4uqdHA3we5jdEXtRl.front131-prd; SGID=sb.springboot131-prd; bm_sz=C3631D7AA73F951FA5D4AD28574EDBAE~YAAQzTVDF/efuJ+IAQAAvrCnrhQzSqHopKHCkEzH4CDBpVSjt2CdcrK7xKD4Ur9r2WMDl2abOsq90ekwmhfXa7QEKxaH1Mk+tSyR4CGeG4q4OCSHFDM9pFimOHLrKfn5EA4a3tCLGxLgLpHfAfL739DE+UQnjZhg69ZeUaZzpRDauJvEKVUVGz6EGAUKGSmW5dSBtI4CCE+p42IFeiRRd/BI0wsYOdwZksOa0NYERiZ6T5u/Rru2R1t6RtpgvIqHB7ByYZsT7fFJJoNVgmWEvWJI24G6lg8qooPgIyXB2P9hGq/FvRsd170=~3552308~3556914; AKA_A2=A; OPTOUTMULTI=0:0%7Cc1:0%7Cc2:0%7Cc4:0%7Cc3:0; lv-dispatch=kor-kr; lv-dispatch-url=https://kr.louisvuitton.com/kor-kr/products/puzzle-flower-monogram-keyring-s00-nvprod4170081v/M01207; ak_cc=KR; ak_bmsc=2F6DAD582D758BD2C3313B28ECC46CCB~000000000000000000000000000000~YAAQzTVDF9OguJ+IAQAA5rWnrhQ3hb2TxLVo5Ex8aajU2QlZT0aikdEl2ssz2cFI0nmQ/GxtYrC3mQocuAQcOCenI5fdJGx8nlN+TdxWnwr/wYMqkV88UK8+gkHKRkw3HdCs7CutvEHjB8w/PdRr+UnXY0qqzsg+nlFxn7n3arUqMiiuP34Cx8+enjzQYWnz0mbqMJFJXJsy+Qc2EmrNuLHj75DEvsNsWwM9Zluteq80yFHayEf6g6C7F/+z2iQZSy92xgotTC2+OA31Iloncux93WZbgRuUHwVWVPyUV0yg2xhBXogJdYY6Ou2TX/l6eftOFnyJzM4Yo0L3adcFohgE9dnTsRY52jHsnSdsYkhVoMJJ6K4UH/uxIUq9lubHZ+iQ3KrNohQN8dk+VIZYQPi40wOo0cu179k3fAB8zxEi7XgzxA5p/66UW6BmA1XtQR3e5dslAddoTp4vZbQ5vRG4hr8SSrtmuCseU97ONf2f; RT="z=1&dm=louisvuitton.com&si=66447985-e7aa-49f7-a5dc-104845852594&ss=liskqdmw&sl=0&tt=0&bcn=%2F%2F684d0d4b.akstat.io%2F"; _cs_mk=0.926590451296496_1686557407779; _fbp=fb.1.1686557408151.960999036; _scid=196b6e5d-818b-4da9-89b4-b5afe7433432; _scid_r=196b6e5d-818b-4da9-89b4-b5afe7433432; _tt_enable_cookie=1; _ttp=i9BoCdSWrv1ulrGKB-ATML_cHrj; _qubitTracker=gkuisyi8co8-0liskqi6w-69tgrt8; qb_generic=:Yiup8Lm:.louisvuitton.com; _gid=GA1.2.1423735359.1686557411; _ga=GA1.1.795833456.1686557411; _gcl_au=1.1.722436059.1686557411; _abck=49A08A3DFA90539A85752575479FAB37~0~YAAQxjVDF4u9OZ+IAQAA/MynrgrzawyY+shhkV6LdQuVG0VXX4iT6zbtcyMsC/Id3nLEQdA1ueCy6lDYrgFj5UoYbINNklZWNY2c1wCjDBwpnz0nCSN08lvY3+oVS7Jiaaq/+0veidDHwvlkz1ODuOyxXZniHRZGh1xiGdp2TzIE8UVXenBS/FzA3ex++VI7KM5DvizNHFO2QXmt9wDmG6yqPc4Rzo4c2EH8WGTx5zurnhQ2q6vEFWPsTvELpzT0lSZqJUz5qW+H+H2fDZkLgpDZwnzqwfTzLjEp+ICBBIt6LahxoINAZL5WTkE+ks2Q+2cgOUMLcsMYO2bUN8t16QCOSPm7yeBcetP4umorik8tHlA+A4ZGdg/BoOG3QUeNJZxMB0Bf49ZH6u4yKrk5lsY6xi9cvOZYo6xI7+s9~-1~-1~-1; qb_permanent=gkuisyi8co8-0liskqi6w-69tgrt8:1:1:1:1:0::0:1:0:BkhtLj:BkhtLj:::::211.52.1.133:seoul:2261:south%20korea:KR:37.56:127:jung-gu%20seoul:410014:seoul-teukbyeolsi:25025:migrated|1686557413582:EbFa==B=CRhL=s&Fv2m==B=CrkC=MQ&F4TI==B=Cs1j=J5::Yiup9VG:Yiup8WB:0:0:0::0:0:.louisvuitton.com:0; qb_session=1:1:32:EbFa=B&Fv2m=B&F4TI=B:0:Yiup8WB:0:0:0:0:.louisvuitton.com; _sctr=1%7C1686495600000; _ga_S6ED35NYJQ=GS1.1.1686557410.1.0.1686557415.55.0.0; utag_main=v_id:0188aea7b40c00231535fc13a6180506f006406700bd0$_sn:1$_se:8$_ss:0$_st:1686559215620$ses_id:1686557406222%3Bexp-session$_pn:1%3Bexp-session$dc_visit:1$dc_event:5%3Bexp-session$dc_region:eu-central-1%3Bexp-session'
                    }
                data = {
                        "flagShip": False,
                        "country": "KR",
                        "query": "서울",
                        "latitudeA": "37.768825744921735",
                        "latitudeB": "37.36061764438869",
                        "latitudeCenter": "37.56472169465521",
                        "longitudeA": "126.51683339179687",
                        "longitudeB": "127.43144520820312",
                        "longitudeCenter": "126.97413929999999",
                        "query": "",
                        "clickAndCollect": False,
                        "skuId": product_code,  
                        "pageType": "productsheet"
                    }
                response = requests.post(url, headers=headers, data=json.dumps(data))
                #print(f"API 응답: {response.text}")  # 로깅
                
                if response.status_code == 200:
                    response_data = response.json()
                    stores_with_stock = [
                        store for store in response_data.get('hits', [])
                        if any(
                            prop.get('value') == 'true' and prop.get('name') == 'stockAvailability'
                            for prop in store.get('additionalProperty', [])
                        )
                    ]
                    total_stores_count = len(stores_with_stock)

                    for store in stores_with_stock:
                        if '루이 비통 서울 도산' in store.get('name') or '루이 비통 메종 서울' in store.get('name'):
                            seoul_dosan_count += 1
                        stock_info.append(store.get('name'))  # 모든 매장 이름 추가

            except Exception as e:
                print(f"재고 조회 중 예외 발생: {e}")

        # 매장 재고 조회 성공 후 처리
        if total_stores_count > 0:
            if max_stores_to_save > 0 and len(stock_info) > max_stores_to_save:
                stock_info_to_save = stock_info[:max_stores_to_save]
            else:
                stock_info_to_save = stock_info
            stock_info_str = '\n'.join(stock_info_to_save)
        else:
            stock_info_str = 'N/A'

        # 재고 현황 문자열 설정
        if total_stores_count == 0 and seoul_dosan_count == 0:
            stock_status = 'N/A'
        else:
            stock_status = f"{total_stores_count}({seoul_dosan_count})"
        
        # 여기에 공홈 등록 상태 업데이트 로직을 추가합니다.
        if price == 'N/A' and stock_info_str == 'N/A':
            website_status = NOT_AVAILABLE
        elif price != 'N/A' and stock_info_str == 'N/A':
            website_status = ON_WEBSITE_NOT_SOLD  
        elif price == 'N/A' and stock_info_str != 'N/A':
            website_status = NOT_ON_WEBSITE_SOLD
        else:
            website_status = ON_WEBSITE_SOLD
            
        # 엑셀 파일에 결과 저장
        sheet.cell(row=row, column=2).value = price
        sheet.cell(row=row, column=3).value = stock_status
        sheet.cell(row=row, column=4).value = website_status  # 수정된 부분
        sheet.cell(row=row, column=5).value = stock_info_str

        # 변경사항을 파일에 저장
        workbook.save(filename=target_path)
        
        # 진행 상태 저장
        save_progress(progress_file, sheet_name, row)
        
        # 콘솔에 결과 출력
        print(f"제품 코드: {product_code}")
        print(f"가격: {price}")
        print(f"전체 매장 수: {total_stores_count}, 서울/도산 매장 수: {seoul_dosan_count}")
        print(f"재고 있는 매장 목록:\n{stock_info_str}")
        
        # 처리된 행 수를 업데이트하고 진행 상황을 출력
        processed_rows += 1
        progress_percentage = (processed_rows / total_rows) * 100  # 진행률 계산
        print(f"현재 작업중인 시트: {sheet_name}, 진행 상태: {processed_rows}/{total_rows} ({progress_percentage:.2f}%)")
        #print(f"현재 작업중인 시트: ({progress_percentage:.2f}%)")
        
        # 다음 제품 코드 검색 전에 대기
        time.sleep(1)
        print("Process completed. File saved successfully.")
        # 제품 코드 처리가 끝날 때마다 콘솔 출력에 공백 줄 추가
        print()
