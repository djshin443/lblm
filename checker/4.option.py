import os
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import time
import traceback

def setup_driver():
    options = Options()
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument('--disable-infobars')
    options.add_argument('--start-maximized')
    options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3')
    options.add_argument('--log-level=3')
    options.add_argument('--disable-features=ChromeWhatsNewUI')
    service = ChromeService(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(10)
    return driver

def load_product_codes(excel_path):
    workbook = load_workbook(filename=excel_path)
    sheet_names = workbook.sheetnames
    product_codes_dict = {}

    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        product_codes = [cell.value for cell in sheet['A'][1:] if cell.value]
        product_codes_dict[sheet_name] = product_codes

    return product_codes_dict

def extract_product_codes(driver):
    product_codes = []
    try:
        # 모든 옵션 확인
        try:
            options = driver.find_elements(By.CSS_SELECTOR, ".lv-product-panel-grid__item .lv-product-card__name")
            for option in options:
                product_id = option.get_attribute("id")
                if product_id and product_id.startswith("product-"):
                    product_code = product_id.replace("product-", "")
                    if product_code not in product_codes:
                        product_codes.append(product_code)
        except Exception as e:
            print(f"Error extracting product codes: {str(e)}")
            traceback.print_exc()

    except Exception as e:
        print(f"Error in main try block: {str(e)}")
        traceback.print_exc()

    return product_codes

def remove_duplicates(sheet):
    all_codes = [cell.value for cell in sheet['A'][1:] if cell.value]
    unique_codes = list(set(all_codes))
    sheet.delete_rows(2, sheet.max_row)
    for row, code in enumerate(unique_codes, start=2):
        sheet.cell(row=row, column=1, value=code)

def save_to_excel(sheet, product_codes, row):
    for code in product_codes:
        sheet.cell(row=row, column=1, value=code)
        row += 1
    return row

def main():
    current_dir = os.getcwd()
    source_filename = '루이비통.xlsx'
    source_path = os.path.join(current_dir, source_filename)
    product_codes_dict = load_product_codes(source_path)
    done_codes = set()
    driver = setup_driver()
    driver.get("https://kr.louisvuitton.com/")

    workbook = load_workbook(filename=source_path)
    for sheet_name, product_codes in product_codes_dict.items():
        sheet = workbook[sheet_name]
        next_row = sheet.max_row + 1
        for product_code in product_codes:
            try:
                if product_code in done_codes:
                    continue

                print(f"Searching for product code: {product_code}")
                WebDriverWait(driver, 20).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, ".lv-header__utility-label.list-label-s"))
                ).click()
                time.sleep(2)
                search_input = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "#searchHeaderInput"))
                )
                search_input.clear()
                search_input.send_keys(product_code)
                search_input.send_keys(Keys.ENTER)
                time.sleep(5)

                try:
                    product_link = WebDriverWait(driver, 20).until(
                        EC.presence_of_element_located((By.XPATH, f"//a[contains(@href, '{product_code}')]"))
                    )
                    product_url = product_link.get_attribute("href")
                    driver.get(product_url)
                    time.sleep(5)

                    extracted_product_codes = extract_product_codes(driver)

                    if extracted_product_codes:
                        next_row = save_to_excel(sheet, extracted_product_codes, next_row)
                        print(f"[품번] {product_code}")
                        print(f"[추출된 품번] {', '.join(extracted_product_codes)}")
                    else:
                        print(f"No options found for product {product_code}")

                    done_codes.add(product_code)

                except Exception as e:
                    print(f"No search results found for product code: {product_code}")
                    print(f"Error: {str(e)}")
                    traceback.print_exc()

                driver.get("https://kr.louisvuitton.com/")
                time.sleep(5)

            except Exception as e:
                print(f"검색 중 오류 발생: {str(e)}")
                traceback.print_exc()
                print("브라우저를 다시 시작합니다.")
                driver.quit()
                driver = setup_driver()
                driver.get("https://kr.louisvuitton.com/")
                continue

        remove_duplicates(sheet)
        workbook.save(filename=source_path)

    driver.quit()
    print("제품 검색 완료")

if __name__ == "__main__":
    main()
