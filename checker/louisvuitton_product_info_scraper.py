import os
import sys
import json
import re
import requests
from datetime import datetime
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from colorama import Fore, Style
import textwrap


def create_initial_file():
    # 파일 이름 정의
    source_filename = '루이비통.xlsx'
    current_dir = os.getcwd()
    source_path = os.path.join(current_dir, source_filename)

    # 파일이 존재하는지 확인
    if os.path.exists(source_path):
        print("루이비통.xlsx 파일이 이미 존재합니다. 계속 진행합니다.")  # 파일이 존재하면 바로 진행
        return  # 사용자 입력 대기를 건너뛰고 바로 종료

    # 파일이 존재하지 않는 경우 생성
    workbook = Workbook()
    sheet = workbook.active
    columns_titles = ['품번', '매장가', '공홈여부', '재고현황', '재고매장', '제품명', '소재', '사이즈', '특징', '세부설명']

    # 열 제목 설정
    for col, title in enumerate(columns_titles, start=1):
        sheet.cell(row=1, column=col, value=title)

    workbook.save(filename=source_path)
    print("루이비통.xlsx 파일을 생성하였습니다. 생성된 파일에 품번을 입력하고 Enter를 눌러 계속 진행하세요.")

    # 사용자 입력 대기 (파일이 새로 생성된 경우에만)
    input()  # 파일 생성 시에만 'Enter'를 대기


if __name__ == '__main__':
    create_initial_file()

# 사용자 정의 상태 변수
ON_WEBSITE_SOLD = "O"  # 공홈에도 있고 실제로 판매 중인 제품
NOT_ON_WEBSITE_SOLD = "X"  # 공홈에는 없지만 실제로 판매 중인 제품
NOT_AVAILABLE = "N/A"  # 공홈에도 없고 실제로 판매하지 않는 제품
ON_WEBSITE_NOT_SOLD = "O"  # 공홈에는 있지만 판매하지 않는 제품
max_stores_to_save = 5

current_dir = os.getcwd()  # 현재 디렉토리 경로를 가져오는 함수입니다.
datetime_now = datetime.now().strftime('%Y%m%d')  # 현재 날짜를 가져오는 함수입니다. (년도월일까지만)
source_filename = '루이비통.xlsx'  # 가져올 엑셀 파일의 이름입니다.
target_filename = f'루이비통_{datetime_now}.xlsx'  # 저장할 엑셀 파일의 이름입니다.
progress_file_name = f'progress_{datetime_now}.txt'  # 진행 상태를 저장할 파일의 이름입니다.

source_path = os.path.join(current_dir, source_filename)  # 가져올 엑셀 파일의 전체 경로입니다.
target_path = os.path.join(current_dir, target_filename)  # 저장할 엑셀 파일의 전체 경로입니다.
progress_file = os.path.join(current_dir, progress_file_name)  # 진행 상태 파일의 전체 경로입니다.


# 진행 상태를 저장하는 함수입니다.
def save_progress(progress_file, sheet_name, row):
    print(f"Calling save_progress with sheet: {sheet_name}, row: {row}")
    try:
        with open(progress_file, 'w') as file:
            file.write(f"{sheet_name},{row}")
    except Exception as e:
        print(f"Failed to save progress: {e}")
    with open(progress_file, 'w') as file:
        file.write(f"{sheet_name},{row}")
    print(f"Saved progress to {progress_file}: {sheet_name},{row}")


def load_progress(progress_file):
    print(f"Loading progress from {progress_file}")
    if os.path.exists(progress_file):
        with open(progress_file, 'r') as file:
            content = file.read().strip()
            print(f"Loaded content: '{content}'")
            if content:
                sheet_name, row = content.split(',')
                print(f"Returning: {sheet_name}, {row}")
                return sheet_name, int(row)
            else:
                print("No content found, returning None, 0")
                return None, 0
    else:
        print("No progress file found, returning None, 0")
        return None, 0


options = Options()
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument('--disable-infobars')
options.add_argument('--headless')
options.add_argument('--window-size=1920,1080')
options.add_argument(
    'user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.3')
options.add_argument('--log-level=3')

service = ChromeService(ChromeDriverManager().install())
service.log_path = os.devnull
driver = webdriver.Chrome(service=service, options=options)
driver.implicitly_wait(5)

workbook = load_workbook(filename=source_path) if os.path.exists(source_path) else Workbook()

# 이미 파일이 존재하는 경우 해당 파일을 로드하고, 존재하지 않는 경우에만 새로 생성합니다.
if os.path.exists(target_path):
    workbook = load_workbook(filename=target_path)
else:
    workbook = Workbook()

    # 기본 시트 제거
    workbook.remove(workbook.active)

    # 기존 파일에서 데이터와 시트 복사
    source_workbook = load_workbook(filename=source_path)
    for sheet_name in source_workbook.sheetnames:
        source_sheet = source_workbook[sheet_name]
        target_sheet = workbook.create_sheet(sheet_name)
        for row in source_sheet.iter_rows(values_only=True):
            target_sheet.append(row)

# 모든 시트에 열 제목 강제 삽입
columns_titles = ['품번', '매장가', '공홈여부', '재고현황', '재고매장', '제품명', '소재', '사이즈', '특징', '세부설명']
for sheet in workbook.worksheets:
    for col, title in enumerate(columns_titles, start=1):
        sheet.cell(row=1, column=col, value=title)

workbook.save(filename=target_path)
last_processed_sheet, last_processed_row = load_progress(progress_file)

    # 엑셀 파일에 정보를 기록하는 로직


def write_to_cell(sheet, column, row_num, value):
    if value is None or value == '' or value == '0':  # '0' 값을 'N/A'로 처리
        value = 'N/A'

    # 엑셀에서 허용하지 않는 문자 제거
    value = ''.join(char for char in value if char.isprintable())

    # 셀에 값 쓰기
    sheet.cell(row=row_num, column=column, value=value)


# 제품 정보 및 매장 가격 가져오기 함수
def get_product_info(product_code):
    try:
        driver.get(f'https://kr.louisvuitton.com/kor-kr/search/{product_code}')
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "body")))

        if "product not found" in driver.page_source.lower():
            raise ValueError("Product not found")

        try:
            price_element = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.CLASS_NAME, 'lv-product__price'))
            )
            price_text = price_element.text.strip()
            price_text = re.sub(r'[₩\s]', '', price_text)
        except Exception as e:
            print(f"가격 정보를 불러오는 중 예외 발생: {e}")
            price_text = "N/A"

        try:
            product_name = driver.find_element(By.CLASS_NAME, 'lv-product__name').text
        except:
            product_name = "N/A"

        try:
            # 소재 정보를 포함하는 버튼 요소를 찾기
            material_button = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "button.lv-product-variation-selector"))
            )
            # '소재'라는 텍스트를 가진 span 요소 찾기
            material_title = material_button.find_element(By.XPATH, ".//span[contains(text(), '소재')]")

            # 소재 값을 가진 span 요소 찾기
            material_value = material_button.find_element(By.CSS_SELECTOR, '.lv-product-variation-selector__value').text
            material = material_value.strip()
        except Exception as e:
            print(f"소재 정보 파싱 중 오류 발생: {e}")
            material = "N/A"

        try:
            details_button = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '.lv-product-features-buttons__button'))
            )
            details_button.click()
            description_text = WebDriverWait(driver, 20).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, '#modalContent p'))
            ).text
        except:
            description_text = "N/A"

        try:
            size_element = driver.find_element(By.CSS_SELECTOR,
                                               '#modalContent > div.lv-product-detailed-features > div > div.lv-product-dimension.body-s.lv-product-detailed-features__dimensions')
            if size_element:
                size_text = size_element.text.strip()
                size_text = ' '.join(size_text.splitlines())
            else:
                size_text = "N/A"
        except Exception as e:
            size_text = "N/A"

        try:
            features = driver.find_elements(By.CSS_SELECTOR, '.lv-product-detailed-features__description li')
            feature_text = "\n".join([feature.text.strip() for feature in features if feature.text.strip() != ""])
        except Exception as e:
            feature_text = "N/A"

        return price_text, product_name, material, size_text, feature_text, description_text

    except Exception as e:
        print(f"제품 {product_code}에 대한 오류 발생: {str(e)}")
        return "N/A", "N/A", "N/A", "N/A", "N/A", "N/A"


# 각 제품 코드에 대한 정보 가져오기
start_row = 2  # 제품 정보를 입력할 시작 행 번호

if last_processed_sheet is None:
    last_processed_sheet = workbook.sheetnames[0]
if last_processed_row is None or last_processed_row < start_row:
    last_processed_row = start_row

print(f"이어서 작업을 시작합니다. 시트: {last_processed_sheet}, 행: {last_processed_row}")

# 모든 시트를 순회하며 처리
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    last_row = sheet.max_row

    if sheet_name == last_processed_sheet:
        start_row = last_processed_row
    else:
        start_row = 2  # 새 시트에서는 처음부터 시작

    for row_num in range(start_row, last_row + 1):
        row_data = list(sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
        if not row_data or not row_data[0]:
            break
        product_code = row_data[0][0]  # row_data[0]은 행 전체, 그 중 첫 번째 열 선택
        if product_code is None:
            break

        product_price, product_name, material, size, features, description = get_product_info(product_code)
        # print(f"제품 코드: {product_code}, 제품명: {product_name}, 가격: {product_price}, 소재: {material}, 사이즈: {size}")
        # print(f"특징: {features}, 설명: {description}")

        # 매장 재고 조회 성공 후 처리
        total_stores_count = 0
        seoul_dosan_count = 0  # 서울/도산 매장 수 초기화
        stock_info = []  # 재고가 있는 매장 이름을 저장할 리스트

        if product_price != 'N/A':  # 가격 정보가 'N/A'가 아닌 경우에만 재고 조회를 실행
            try:
                # 매장 재고 조회
                total_stores_count = 0
                seoul_dosan_count = 0  # 서울/도산 매장 수 초기화
                stock_info = []  # 재고가 있는 매장 이름을 저장할 리스트

                if product_price != 'N/A':  # 가격 정보가 'N/A'가 아닌 경우에만 재고 조회를 실행
                    try:
                        url = 'https://api.louisvuitton.com/eco-eu/search-merch-eapi/v1/kor-kr/stores/query'
                        headers = {
                            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36',
                            'Content-Type': 'application/json',
                            'Accept': 'application/json, text/plain, */*',
                            'client_secret': '60bbcdcD722D411B88cBb72C8246a22F',
                            'client_id': '607e3016889f431fb8020693311016c9',
                            'Origin': 'https://kr.louisvuitton.com',
                            'Accept-Language': 'ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7',
                        }
                        data = {
                            "flagShip": False,
                            "country": "KR",
                            "query": "서울",
                            "latitudeA": "37.768825744921735",
                            "latitudeB": "37.36061764438869",
                            "latitudeCenter": "37.56472169465521",
                            "longitudeA": "126.51683339179687",
                            "longitudeB": "127.43144520820312",
                            "longitudeCenter": "126.97413929999999",
                            "query": "",
                            "clickAndCollect": False,
                            "skuId": product_code,
                            "pageType": "productsheet"
                        }
                        response = requests.post(url, headers=headers, data=json.dumps(data))
                        # print(f"API 응답: {response.text}")  # 로깅

                        if response.status_code == 200:
                            response_data = response.json()
                            stores_with_stock = [
                                store for store in response_data.get('hits', [])
                                if any(
                                    prop.get('value') == 'true' and prop.get('name') == 'stockAvailability'
                                    for prop in store.get('additionalProperty', [])
                                )
                            ]
                            total_stores_count = len(stores_with_stock)

                            for store in stores_with_stock:
                                store_name = store.get('name')
                                stock_info.append(store_name)

                            # 메종 서울/ 서울 도산 매장 카운트
                            if '루이 비통 메종 서울' in store_name or '루이 비통 서울 도산' in store_name:
                                seoul_dosan_count += 1
                        else:
                            print(f"API 요청이 실패했습니다. 상태 코드: {response.status_code}")
                    except Exception as e:
                        print(f"재고 조회 중 오류 발생: {e}")
                        continue

                # 재고 정보 출력
                # print(f"총 {total_stores_count} 개 매장에서 재고를 찾았습니다.")
                # print(f"서울/도산 매장 중 {seoul_dosan_count} 개 매장에서 재고를 찾았습니다.")
                # print("재고 있는 매장:", stock_info)

                # 매장 재고 조회 성공 후 처리
                if total_stores_count > 0:
                    if max_stores_to_save > 0 and len(stock_info) > max_stores_to_save:
                        stock_info_to_save = stock_info[:max_stores_to_save]
                    else:
                        stock_info_to_save = stock_info
                    stock_info_str = '\n'.join(stock_info_to_save)
                else:
                    stock_info_str = 'N/A'

                # 재고 현황 문자열 설정
                if total_stores_count == 0 and seoul_dosan_count == 0:
                    stock_status = 'N/A'
                else:
                    stock_status = f"{total_stores_count}({seoul_dosan_count})"

                # 공홈 등록 상태 업데이트
                if product_price == 'N/A' and stock_info_str == 'N/A':
                    website_status = NOT_AVAILABLE
                elif product_price != 'N/A' and stock_info_str == 'N/A':
                    website_status = ON_WEBSITE_NOT_SOLD
                elif product_price == 'N/A' and stock_info_str != 'N/A':
                    website_status = NOT_ON_WEBSITE_SOLD
                else:
                    website_status = ON_WEBSITE_SOLD

                # 각 셀에 값을 쓰는 부분을 'write_to_cell' 함수를 사용해 대체합니다.
                write_to_cell(sheet, 2, row_num, product_price)  # 매장가
                write_to_cell(sheet, 3, row_num, website_status)  # 공홈여부
                write_to_cell(sheet, 4, row_num, stock_status)  # 재고현황
                write_to_cell(sheet, 5, row_num, stock_info_str)  # 재고매장
                write_to_cell(sheet, 6, row_num, product_name)  # 제품명
                write_to_cell(sheet, 7, row_num, material)  # 소재
                write_to_cell(sheet, 8, row_num, size)  # 사이즈
                write_to_cell(sheet, 9, row_num, features)  # 특징
                write_to_cell(sheet, 10, row_num, description)  # 세부 설명


                # 제품 정보 출력
                def print_product_info(product_code, product_price, website_status, stock_status, stock_info,
                                       product_name, material, size, features, description):
                    # 문자열을 주어진 너비에 맞춰 왼쪽으로 정렬하고 줄바꿈
                    def format_text_block(text, width=70):
                        return textwrap.fill(text, width=width)

                    # 세부 설명 포맷팅
                    formatted_description = format_text_block(description)
                    formatted_features = format_text_block(features)

                    # 재고 매장 정보 포맷팅: 매장명을 반복하며 새로운 줄에 출력
                    formatted_stock_info = "\n".join(stock_info)

                    print(Fore.LIGHTCYAN_EX + "[제품 정보]:" + Style.RESET_ALL)
                    print(Fore.LIGHTMAGENTA_EX + f"품번: {product_code}" + Style.RESET_ALL)
                    print(Fore.LIGHTGREEN_EX + f"매장가: {product_price}" + Style.RESET_ALL)
                    print(Fore.LIGHTYELLOW_EX + f"공홈여부: {website_status}" + Style.RESET_ALL)
                    print(Fore.LIGHTBLUE_EX + f"재고현황: {stock_status}" + Style.RESET_ALL)
                    print(Fore.LIGHTMAGENTA_EX + f"재고매장:\n{formatted_stock_info}" + Style.RESET_ALL)
                    print(Fore.LIGHTRED_EX + f"제품명: {product_name}" + Style.RESET_ALL)
                    print(Fore.LIGHTCYAN_EX + f"소재: {material}" + Style.RESET_ALL)
                    print(Fore.LIGHTGREEN_EX + f"사이즈: {size}" + Style.RESET_ALL)
                    print(Fore.LIGHTYELLOW_EX + f"특징:\n{formatted_features}" + Style.RESET_ALL)
                    print(Fore.LIGHTBLUE_EX + f"세부 설명:\n{formatted_description}" + Style.RESET_ALL)


                # 이 함수를 호출하는 부분에서는 위 함수를 활용해 제품 정보를 출력합니다.
                print_product_info(product_code, product_price, website_status, stock_status, stock_info, product_name,
                                   material, size, features, description)

                # 저장
                workbook.save(filename=target_path)
                save_progress(progress_file, sheet_name, row_num)
                print(f"제품 정보를 엑셀 파일에 저장했습니다.")

            except Exception as e:
                print(f"재고 조회 중 오류 발생: {e}")
                continue
    last_processed_row = start_row

    # 현재 시트 처리 완료 후, 다음 시트를 위해 last_processed_row 초기화
    last_processed_row = 2  # 다음 시트를 위해 초기화

# 작업 완료 메시지 출력
print("작업이 완료되었습니다.")
